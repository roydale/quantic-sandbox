import openpyxl
from openpyxl.styles import PatternFill
from analyze_data import read_sales_data, compare_product_sales
import os

def highlight_sales_difference(excel_file_path):
    # Read the sales data from the Excel files using the analyze_data functions
    sales_data = read_sales_data(excel_file_path)
    comparison_result = compare_product_sales(sales_data)

    # Loop through each Excel file and highlight cells with differing sales in yellow
    for file_name, df in sales_data.items():
        # Construct the full file path using os.path.join
        file_path = os.path.join(excel_file_path, file_name)

        # Open the Excel file using openpyxl
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Initialize a row counter
        current_row = 2  # Start from row 2 (assuming the header is in row 1)

        # Get the column indices for 'Product' and 'Sold'
        product_column = None
        sold_column = None
        for col in sheet.iter_cols(min_row=1, max_row=1):
            for cell in col:
                if cell.value == 'Product':
                    product_column = col[0].column
                elif cell.value == 'Sold':
                    sold_column = col[0].column

        # Iterate through the rows and highlight cells with differing sales in yellow
        for row in sheet.iter_rows(min_row=2, values_only=True):
            product = row[product_column - 1]
            sold = row[sold_column - 1]

            if product in comparison_result and file_name in comparison_result[product]['varying_files']:
                # Highlight the 'Sold' cell in yellow
                sold_cell = sheet.cell(row=current_row, column=sold_column)
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                sold_cell.fill = yellow_fill

            # Increment the row counter
            current_row += 1

        # Save the modified Excel file
        workbook.save(file_path)

if __name__ == "__main__":
    directory_path = "sales"  # Replace with the path to your sales directory
    highlight_sales_difference(directory_path)
